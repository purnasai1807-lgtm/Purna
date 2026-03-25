from __future__ import annotations

import re
from collections.abc import Iterator
from typing import IO, Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pandas.api.types import is_datetime64_any_dtype, is_numeric_dtype
from sklearn.linear_model import LinearRegression

from app.services.modeling import build_modeling_summary
from app.services.visualization import generate_chart_specs

TYPE_INFERENCE_SAMPLE_SIZE = 2000
CORRELATION_SAMPLE_SIZE = 15000
UPLOAD_READ_CHUNK_SIZE = 5000


def parse_uploaded_dataframe(filename: str, upload_stream: IO[bytes]) -> pd.DataFrame:
    extension = filename.lower().rsplit(".", maxsplit=1)[-1] if "." in filename else ""
    upload_stream.seek(0)

    if extension == "csv":
        return read_csv_in_chunks(upload_stream)
    if extension in {"xlsx", "xlsm", "xltx", "xltm"}:
        return read_excel_in_chunks(upload_stream)
    if extension == "xls":
        return read_legacy_excel_in_chunks(upload_stream)
    if extension == "json":
        try:
            return pd.read_json(upload_stream)
        except ValueError as exc:
            raise ValueError("The uploaded JSON file could not be parsed.") from exc

    raise ValueError("Unsupported file type. Please upload a CSV, Excel, or JSON file.")


def read_csv_in_chunks(upload_stream: IO[bytes]) -> pd.DataFrame:
    chunks: list[pd.DataFrame] = []
    try:
        reader = pd.read_csv(
            upload_stream,
            chunksize=UPLOAD_READ_CHUNK_SIZE,
            low_memory=False,
        )
    except pd.errors.EmptyDataError as exc:
        raise ValueError("The uploaded file is empty.") from exc

    for chunk in reader:
        chunks.append(chunk)

    if chunks:
        return pd.concat(chunks, ignore_index=True)

    upload_stream.seek(0)
    return pd.read_csv(upload_stream, nrows=0)


def read_excel_in_chunks(upload_stream: IO[bytes]) -> pd.DataFrame:
    workbook = load_workbook(upload_stream, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            raise ValueError("The uploaded file is empty.")

        columns = build_excel_columns(header_row)
        return build_dataframe_from_excel_rows(rows, columns)
    finally:
        workbook.close()


def read_legacy_excel_in_chunks(upload_stream: IO[bytes]) -> pd.DataFrame:
    chunks: list[pd.DataFrame] = []
    rows_to_skip = 0

    while True:
        upload_stream.seek(0)
        chunk = pd.read_excel(
            upload_stream,
            nrows=UPLOAD_READ_CHUNK_SIZE,
            skiprows=(
                lambda row_index, offset=rows_to_skip: 0 < row_index <= offset
            )
            if rows_to_skip
            else None,
        )

        if chunk.empty:
            if chunks:
                break
            return chunk

        chunks.append(chunk)
        rows_to_skip += len(chunk)
        if len(chunk) < UPLOAD_READ_CHUNK_SIZE:
            break

    return pd.concat(chunks, ignore_index=True)


def build_excel_columns(header_row: tuple[Any, ...]) -> list[str]:
    columns: list[str] = []
    for index, value in enumerate(header_row):
        if value is None or (isinstance(value, str) and not value.strip()):
            columns.append(f"Unnamed: {index}")
        else:
            columns.append(str(value))
    return columns


def build_dataframe_from_excel_rows(
    rows: Iterator[tuple[Any, ...]],
    columns: list[str],
) -> pd.DataFrame:
    chunks: list[pd.DataFrame] = []
    chunk_rows: list[list[Any]] = []

    for row in rows:
        padded_row = list(row[: len(columns)])
        if len(padded_row) < len(columns):
            padded_row.extend([None] * (len(columns) - len(padded_row)))
        chunk_rows.append(padded_row)

        if len(chunk_rows) >= UPLOAD_READ_CHUNK_SIZE:
            chunks.append(pd.DataFrame(chunk_rows, columns=columns))
            chunk_rows = []

    if chunk_rows:
        chunks.append(pd.DataFrame(chunk_rows, columns=columns))

    if chunks:
        return pd.concat(chunks, ignore_index=True)

    return pd.DataFrame(columns=columns)


def parse_manual_dataframe(
    columns: list[str],
    rows: list[dict[str, Any]],
) -> pd.DataFrame:
    if not rows:
        raise ValueError("Manual entry needs at least one populated row.")

    dataframe = pd.DataFrame(rows)
    if columns:
        missing_columns = [column for column in columns if column not in dataframe.columns]
        for column in missing_columns:
            dataframe[column] = None
        dataframe = dataframe[columns]

    return dataframe


def analyze_dataframe(
    dataframe: pd.DataFrame,
    dataset_name: str,
    source_type: str,
    target_column: str | None = None,
    *,
    include_charts: bool = True,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if dataframe.empty:
        raise ValueError(
            "The dataset is empty after parsing. Please upload a file with records."
        )

    original_dataframe = dataframe.copy()
    cleaned_dataframe, cleaning_summary = clean_dataframe(dataframe.copy())
    if cleaned_dataframe.empty:
        raise ValueError(
            "The dataset is empty after cleaning. Please provide rows with actual values."
        )

    normalized_target = normalize_column_name(target_column) if target_column else None
    if normalized_target and normalized_target not in cleaned_dataframe.columns:
        raise ValueError(
            f"Target column '{target_column}' was not found after cleaning. "
            "Use one of the cleaned column names shown in the dashboard."
        )

    overview = build_overview(
        original_dataframe,
        cleaned_dataframe,
        cleaning_summary,
        normalized_target,
    )
    summary_statistics = build_summary_statistics(cleaned_dataframe)
    correlations = build_correlation_analysis(cleaned_dataframe)
    outliers = detect_outliers(cleaned_dataframe)
    trends = analyze_trends(cleaned_dataframe)
    charts = generate_chart_specs(cleaned_dataframe, correlations) if include_charts else []
    modeling = build_modeling_summary(cleaned_dataframe, normalized_target)
    insights, recommendations = build_narrative(
        overview=overview,
        cleaning=cleaning_summary,
        correlations=correlations,
        outliers=outliers,
        trends=trends,
        modeling=modeling,
    )

    payload = {
        "dataset_name": dataset_name,
        "source_type": source_type,
        "target_column": normalized_target,
        "overview": overview,
        "cleaning": cleaning_summary,
        "summary_statistics": summary_statistics,
        "correlations": correlations,
        "outliers": outliers,
        "trends": trends,
        "charts": charts,
        "modeling": modeling,
        "insights": insights,
        "recommendations": recommendations,
        "metadata": {
            "is_preview": False,
            "processing_mode": "instant",
            "file_type": source_type,
            "file_size_bytes": None,
            "sample_row_count": int(len(cleaned_dataframe)),
            **(metadata or {}),
        },
        "sections": build_section_status(charts_ready=include_charts),
    }
    return sanitize_for_json(payload)


def build_section_status(*, charts_ready: bool) -> dict[str, bool]:
    return {
        "overview": True,
        "summary_statistics": True,
        "correlations": True,
        "outliers": True,
        "trends": True,
        "charts": charts_ready,
        "rows": True,
        "modeling": True,
        "insights": True,
        "recommendations": True,
    }


def clean_dataframe(dataframe: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    original_shape = dataframe.shape
    original_columns = [str(column) for column in dataframe.columns]
    standardized_columns = make_unique(
        [normalize_column_name(column) for column in dataframe.columns]
    )
    dataframe.columns = standardized_columns

    empty_rows_dropped = int(dataframe.isna().all(axis=1).sum())
    dataframe = dataframe.dropna(how="all")

    all_null_columns = [
        column for column in dataframe.columns if dataframe[column].isna().all()
    ]
    dataframe = dataframe.drop(columns=all_null_columns)

    for column in dataframe.columns:
        if dataframe[column].dtype == object:
            dataframe[column] = dataframe[column].map(
                lambda value: value.strip() if isinstance(value, str) else value
            )

    dataframe = infer_column_types(dataframe)

    missing_before = int(dataframe.isna().sum().sum())
    duplicate_rows_removed = int(dataframe.duplicated().sum())
    dataframe = dataframe.drop_duplicates()

    for column in dataframe.columns:
        series = dataframe[column]
        if is_numeric_dtype(series):
            fill_value = series.median()
            if pd.notna(fill_value):
                dataframe[column] = series.fillna(fill_value)
        elif is_datetime64_any_dtype(series):
            if series.notna().any():
                dataframe[column] = series.ffill().bfill()
        else:
            mode = series.mode(dropna=True)
            fill_value = mode.iloc[0] if not mode.empty else "Unknown"
            dataframe[column] = series.fillna(fill_value)

    missing_after = int(dataframe.isna().sum().sum())
    type_map = {column: str(dtype) for column, dtype in dataframe.dtypes.items()}

    return dataframe, {
        "original_shape": {"rows": int(original_shape[0]), "columns": int(original_shape[1])},
        "cleaned_shape": {"rows": int(dataframe.shape[0]), "columns": int(dataframe.shape[1])},
        "column_mapping": dict(zip(original_columns, standardized_columns)),
        "removed_all_null_columns": all_null_columns,
        "empty_rows_dropped": empty_rows_dropped,
        "duplicate_rows_removed": duplicate_rows_removed,
        "missing_values_before": missing_before,
        "missing_values_after": missing_after,
        "detected_data_types": type_map,
    }


def infer_column_types(dataframe: pd.DataFrame) -> pd.DataFrame:
    for column in dataframe.columns:
        series = dataframe[column]
        if is_numeric_dtype(series) or is_datetime64_any_dtype(series):
            continue

        non_null = series.dropna()
        if non_null.empty:
            continue

        sampled_values = (
            non_null.sample(n=TYPE_INFERENCE_SAMPLE_SIZE, random_state=42)
            if len(non_null) > TYPE_INFERENCE_SAMPLE_SIZE
            else non_null
        )
        text_values = sampled_values.astype(str)
        datetime_candidate = pd.to_datetime(text_values, errors="coerce")
        has_datetime_pattern = text_values.str.contains(r"[-/:T]", regex=True).mean() >= 0.5
        if has_datetime_pattern and datetime_candidate.notna().mean() >= 0.85:
            dataframe[column] = pd.to_datetime(series, errors="coerce")
            continue

        numeric_candidate = pd.to_numeric(text_values, errors="coerce")
        if numeric_candidate.notna().mean() >= 0.85:
            dataframe[column] = pd.to_numeric(series, errors="coerce")

    return dataframe


def build_overview(
    original_dataframe: pd.DataFrame,
    cleaned_dataframe: pd.DataFrame,
    cleaning_summary: dict[str, Any],
    target_column: str | None,
) -> dict[str, Any]:
    profiles = []
    for column in cleaned_dataframe.columns:
        series = cleaned_dataframe[column]
        non_null = series.dropna()
        sample_values = [sanitize_for_json(value) for value in non_null.head(3).tolist()]
        profiles.append(
            {
                "column": column,
                "dtype": str(series.dtype),
                "missing_count": int(series.isna().sum()),
                "missing_percentage": round(float(series.isna().mean() * 100), 2),
                "unique_values": int(series.nunique(dropna=True)),
                "sample_values": sample_values,
            }
        )

    return {
        "row_count": int(cleaned_dataframe.shape[0]),
        "column_count": int(cleaned_dataframe.shape[1]),
        "original_row_count": int(original_dataframe.shape[0]),
        "original_column_count": int(original_dataframe.shape[1]),
        "target_column": target_column,
        "preview_rows": sanitize_for_json(
            cleaned_dataframe.head(12).to_dict(orient="records")
        ),
        "columns": profiles,
        "detected_data_types": cleaning_summary.get("detected_data_types", {}),
    }


def build_summary_statistics(dataframe: pd.DataFrame) -> list[dict[str, Any]]:
    statistics: list[dict[str, Any]] = []
    for column in dataframe.columns:
        series = dataframe[column]
        base_stats: dict[str, Any] = {
            "column": column,
            "dtype": str(series.dtype),
            "non_null_count": int(series.notna().sum()),
            "unique_values": int(series.nunique(dropna=True)),
        }

        if is_numeric_dtype(series):
            base_stats.update(
                {
                    "mean": round(float(series.mean()), 4),
                    "median": round(float(series.median()), 4),
                    "std": round(float(series.std(ddof=0)), 4),
                    "min": round(float(series.min()), 4),
                    "max": round(float(series.max()), 4),
                    "q1": round(float(series.quantile(0.25)), 4),
                    "q3": round(float(series.quantile(0.75)), 4),
                }
            )
        elif is_datetime64_any_dtype(series):
            base_stats.update(
                {
                    "min": sanitize_for_json(series.min()),
                    "max": sanitize_for_json(series.max()),
                }
            )
        else:
            value_counts = series.astype(str).value_counts()
            top_value = value_counts.index[0] if not value_counts.empty else None
            top_frequency = int(value_counts.iloc[0]) if not value_counts.empty else 0
            base_stats.update(
                {
                    "top_value": top_value,
                    "top_frequency": top_frequency,
                }
            )

        statistics.append(sanitize_for_json(base_stats))

    return statistics


def build_correlation_analysis(dataframe: pd.DataFrame) -> dict[str, Any]:
    numeric_dataframe = dataframe.select_dtypes(include=[np.number])
    if len(numeric_dataframe) > CORRELATION_SAMPLE_SIZE:
        numeric_dataframe = numeric_dataframe.sample(
            n=CORRELATION_SAMPLE_SIZE,
            random_state=42,
        )

    if numeric_dataframe.shape[1] < 2:
        return {"available": False, "columns": [], "matrix": [], "strongest_pairs": []}

    correlation_matrix = numeric_dataframe.corr().round(3)
    strongest_pairs: list[dict[str, Any]] = []
    columns = list(correlation_matrix.columns)
    for index, left_column in enumerate(columns):
        for right_column in columns[index + 1 :]:
            strongest_pairs.append(
                {
                    "left_column": left_column,
                    "right_column": right_column,
                    "correlation": round(
                        float(correlation_matrix.loc[left_column, right_column]),
                        3,
                    ),
                }
            )

    strongest_pairs = sorted(
        strongest_pairs,
        key=lambda pair: abs(pair["correlation"]),
        reverse=True,
    )[:6]

    return {
        "available": True,
        "columns": columns,
        "matrix": sanitize_for_json(correlation_matrix.values.tolist()),
        "strongest_pairs": strongest_pairs,
    }


def detect_outliers(dataframe: pd.DataFrame) -> list[dict[str, Any]]:
    outlier_results: list[dict[str, Any]] = []
    for column in dataframe.select_dtypes(include=[np.number]).columns:
        series = dataframe[column].dropna()
        if len(series) < 4:
            continue

        q1 = series.quantile(0.25)
        q3 = series.quantile(0.75)
        iqr = q3 - q1
        if iqr == 0:
            continue

        lower_bound = q1 - (1.5 * iqr)
        upper_bound = q3 + (1.5 * iqr)
        outlier_mask = (series < lower_bound) | (series > upper_bound)
        count = int(outlier_mask.sum())
        if count:
            outlier_results.append(
                {
                    "column": column,
                    "count": count,
                    "percentage": round(float(count / len(series) * 100), 2),
                    "lower_bound": round(float(lower_bound), 4),
                    "upper_bound": round(float(upper_bound), 4),
                }
            )

    return outlier_results


def analyze_trends(dataframe: pd.DataFrame) -> list[dict[str, Any]]:
    numeric_columns = list(dataframe.select_dtypes(include=[np.number]).columns)
    if not numeric_columns:
        return []

    datetime_columns = [
        column for column in dataframe.columns if is_datetime64_any_dtype(dataframe[column])
    ]
    ordered_dataframe = (
        dataframe.sort_values(datetime_columns[0])
        if datetime_columns
        else dataframe.copy()
    )
    x_values = np.arange(len(ordered_dataframe)).reshape(-1, 1)
    basis = datetime_columns[0] if datetime_columns else "row_order"
    trend_results: list[dict[str, Any]] = []

    for column in numeric_columns[:4]:
        y_values = ordered_dataframe[column].to_numpy(dtype=float)
        if len(np.unique(y_values)) < 2:
            trend_results.append(
                {
                    "column": column,
                    "direction": "stable",
                    "basis": basis,
                    "description": (
                        f"{column} stays mostly stable across the available records."
                    ),
                }
            )
            continue

        model = LinearRegression()
        model.fit(x_values, y_values)
        slope = float(model.coef_[0])
        y_range = float(np.ptp(y_values)) or 1.0
        normalized_slope = slope * max(len(y_values) - 1, 1) / y_range
        if normalized_slope > 0.08:
            direction = "upward"
        elif normalized_slope < -0.08:
            direction = "downward"
        else:
            direction = "stable"

        trend_results.append(
            {
                "column": column,
                "direction": direction,
                "basis": basis,
                "slope": round(slope, 4),
                "description": (
                    f"{column} shows an {direction} trend when records are ordered by {basis}."
                    if basis != "row_order"
                    else f"{column} shows an {direction} trend across the dataset order."
                ),
            }
        )

    return trend_results


def build_narrative(
    overview: dict[str, Any],
    cleaning: dict[str, Any],
    correlations: dict[str, Any],
    outliers: list[dict[str, Any]],
    trends: list[dict[str, Any]],
    modeling: dict[str, Any],
) -> tuple[list[str], list[str]]:
    insights: list[str] = []
    recommendations: list[str] = []

    insights.append(
        f"The dataset was cleaned from {overview['original_row_count']} rows to "
        f"{overview['row_count']} rows and now contains "
        f"{overview['column_count']} usable columns."
    )

    if cleaning.get("duplicate_rows_removed"):
        insights.append(
            f"{cleaning['duplicate_rows_removed']} duplicate rows were removed before analysis."
        )

    missing_before = cleaning.get("missing_values_before", 0)
    missing_after = cleaning.get("missing_values_after", 0)
    if missing_before:
        insights.append(
            f"The pipeline handled {missing_before - missing_after} missing values automatically."
        )
        recommendations.append(
            "Review the columns with imputed values and strengthen data collection "
            "rules upstream."
        )

    strongest_pairs = correlations.get("strongest_pairs", [])
    if strongest_pairs:
        strongest = strongest_pairs[0]
        insights.append(
            f"{strongest['left_column']} and {strongest['right_column']} move "
            f"together with a correlation of {strongest['correlation']}."
        )
        recommendations.append(
            "Use the strongest correlated features when prioritizing KPI drivers "
            "or dashboard alerts."
        )

    if outliers:
        highest_outlier = sorted(
            outliers,
            key=lambda item: item["count"],
            reverse=True,
        )[0]
        insights.append(
            f"{highest_outlier['column']} contains {highest_outlier['count']} "
            f"potential outliers ({highest_outlier['percentage']}% of records)."
        )
        recommendations.append(
            "Inspect the largest outlier columns to confirm whether extreme values "
            "are real events or data quality issues."
        )

    if trends:
        insights.append(trends[0]["description"])

    if modeling.get("status") == "completed":
        mode = modeling.get("mode")
        selected_model = modeling.get("selected_model", "baseline model")
        metrics = modeling.get("metrics", {})
        if mode == "regression":
            insights.append(
                f"{selected_model} achieved an R2 score of {metrics.get('r2', 'n/a')} "
                "on the held-out test split."
            )
            recommendations.append(
                "If you need stronger forecasts, add more business context columns "
                "and compare the recommended regressors."
            )
        elif mode == "classification":
            insights.append(
                f"{selected_model} reached {metrics.get('accuracy', 'n/a')} accuracy "
                "on the held-out test split."
            )
            recommendations.append(
                "Validate class imbalance and consider threshold tuning before "
                "operationalizing predictions."
            )
        else:
            insights.append(
                f"KMeans found {len(modeling.get('cluster_summary', {}))} segments "
                f"with a silhouette score of {metrics.get('silhouette_score', 'n/a')}."
            )
            recommendations.append(
                "Review cluster segments with domain experts and label them for "
                "sales, customer, or operational actions."
            )
    else:
        recommendations.append(
            "Provide a target column when you want supervised model evaluation for "
            "forecasting or classification."
        )

    if not recommendations:
        recommendations.append(
            "The dataset is analysis-ready. Review the charts and model "
            "recommendations for next steps."
        )

    return insights[:6], recommendations[:6]


def sanitize_for_json(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): sanitize_for_json(item) for key, item in value.items()}
    if isinstance(value, list):
        return [sanitize_for_json(item) for item in value]
    if isinstance(value, tuple):
        return [sanitize_for_json(item) for item in value]
    if isinstance(value, pd.DataFrame):
        return sanitize_for_json(value.to_dict(orient="records"))
    if isinstance(value, pd.Series):
        return sanitize_for_json(value.tolist())
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        if pd.isna(value):
            return None
        return round(float(value), 4)
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, (pd.Timestamp,)):
        return value.isoformat()
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except TypeError:
            pass
    if pd.isna(value):
        return None
    return value


def normalize_column_name(value: Any) -> str:
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "column"


def make_unique(columns: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    unique_columns: list[str] = []

    for column in columns:
        count = counts.get(column, 0)
        if count == 0:
            unique_columns.append(column)
        else:
            unique_columns.append(f"{column}_{count + 1}")
        counts[column] = count + 1

    return unique_columns
