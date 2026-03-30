from __future__ import annotations
import copy
import logging
import random
from datetime import datetime, timedelta, timezone
from dataclasses import dataclass
from pathlib import Path
from typing import Any
import duckdb
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload
from app.core.config import settings
from app.db.models import AnalysisCacheEntry, AnalysisReport, AnalysisReportCacheLink, User
from app.db.session import SessionLocal
from app.services.analytics import (
    analyze_dataframe,
    analyze_trends,
    build_correlation_analysis,
    build_excel_columns,
    build_narrative,
    build_overview,
    build_section_status,
    clean_dataframe,
    normalize_column_name,
    parse_uploaded_dataframe,
    sanitize_for_json,
)
from app.services.modeling import build_modeling_summary
from app.services.storage import StoredUpload, build_parquet_path
from app.services.storage import ensure_local_storage_copy
from app.services.visualization import (
    DASHBOARD_COLORS,
    HEATMAP_SCALE,
    build_chart_payload,
    generate_chart_specs,
    style_figure,
)
MAX_CORRELATION_COLUMNS = 8
MAX_EXACT_CHART_POINTS = 180
MAX_EXACT_SCATTER_BINS = 28
PROCESSING_STATUSES = {"queued", "preview_ready", "processing"}
FINAL_STATUSES = {"completed", "failed"}
SUPPORTED_REPORT_SECTIONS = {
    "overview",
    "summary_statistics",
    "correlations",
    "outliers",
    "trends",
    "charts",
    "modeling",
    "insights",
    "recommendations",
}
logger = logging.getLogger(__name__)
SAMPLING_RANDOM_SEED = 42
def utcnow() -> datetime:
    return datetime.now(timezone.utc)
@dataclass(slots=True)
class ColumnDefinition:
    original_name: str
    normalized_name: str
    inferred_kind: str
    display_dtype: str
def build_processing_strategy(processing_mode: str) -> str:
    if processing_mode == "small":
        return "direct"
    if processing_mode == "medium":
        return "chunked"
    return "optimized_background"
def is_optimized_mode(processing_mode: str) -> bool:
    return processing_mode == "large"
def get_preview_ready_state(processing_mode: str) -> tuple[int, str]:
    if processing_mode == "small":
        return 100, "Analytics complete."
    if processing_mode == "medium":
        return 30, "Chunk-based analytics are running in the background."
    return 24, "Large dataset detected. Processing in optimized mode."
def find_cache_entry(
    db: Session,
    *,
    content_hash: str,
    target_column: str | None,
) -> AnalysisCacheEntry | None:
    statement = select(AnalysisCacheEntry).where(
        AnalysisCacheEntry.content_hash == content_hash,
        AnalysisCacheEntry.target_column == target_column,
    )
    return db.scalar(statement)
def create_cache_entry(
    db: Session,
    *,
    stored_upload: StoredUpload,
    target_column: str | None,
    preview_payload: dict[str, Any],
) -> AnalysisCacheEntry:
    cache_entry = AnalysisCacheEntry(
        content_hash=stored_upload.content_hash,
        original_filename=stored_upload.original_filename,
        file_type=stored_upload.file_type,
        target_column=target_column,
        processing_mode=stored_upload.processing_mode,
        status="completed" if stored_upload.processing_mode == "small" else "preview_ready",
        progress=get_preview_ready_state(stored_upload.processing_mode)[0],
        progress_message=get_preview_ready_state(stored_upload.processing_mode)[1],
        file_size_bytes=stored_upload.file_size_bytes,
        row_count=preview_payload.get("overview", {}).get("row_count", 0),
        column_count=preview_payload.get("overview", {}).get("column_count", 0),
        storage_backend=stored_upload.storage_backend,
        storage_key=stored_upload.storage_key,
        storage_path=str(stored_upload.storage_path),
        preview_payload=preview_payload,
        full_payload=preview_payload if stored_upload.processing_mode == "small" else None,
        sections_ready=preview_payload.get("sections", build_section_status(charts_ready=False)),
    )
    db.add(cache_entry)
    db.flush()
    return cache_entry


def create_upload_report(
    db: Session,
    *,
    current_user: User,
    dataset_name: str,
    target_column: str | None,
    cache_entry: AnalysisCacheEntry,
    source_type: str = "upload",
) -> AnalysisReport:
    report = AnalysisReport(
        user_id=current_user.id,
        dataset_name=dataset_name,
        source_type=source_type,
        target_column=target_column,
        status=cache_entry.status,
        row_count=cache_entry.row_count,
        column_count=cache_entry.column_count,
        report_payload={},
        notes=cache_entry.progress_message,
    )
    db.add(report)
    db.flush()
    db.add(AnalysisReportCacheLink(report_id=report.id, cache_entry_id=cache_entry.id))
    sync_report_from_cache(report, cache_entry)
    db.flush()
    return report


def attach_report_to_existing_cache(
    db: Session,
    *,
    current_user: User,
    dataset_name: str,
    target_column: str | None,
    cache_entry: AnalysisCacheEntry,
) -> AnalysisReport:
    report = AnalysisReport(
        user_id=current_user.id,
        dataset_name=dataset_name,
        source_type="upload",
        target_column=target_column,
        status=cache_entry.status,
        row_count=cache_entry.row_count,
        column_count=cache_entry.column_count,
        report_payload={},
        notes="Loaded from cached analytics." if cache_entry.status == "completed" else cache_entry.progress_message,
    )
    db.add(report)
    db.flush()
    db.add(AnalysisReportCacheLink(report_id=report.id, cache_entry_id=cache_entry.id))
    sync_report_from_cache(report, cache_entry, from_cache=True)
    db.flush()
    return report


def materialize_report_payload(
    report: AnalysisReport,
    cache_entry: AnalysisCacheEntry | None,
    *,
    from_cache: bool = False,
) -> dict[str, Any]:
    if not cache_entry:
        payload = copy.deepcopy(report.report_payload or {})
        payload.setdefault("metadata", {})
        payload["metadata"].setdefault("is_preview", False)
        payload["metadata"].setdefault("processing_mode", report.source_type)
        payload["metadata"].setdefault("optimized_mode", False)
        payload["metadata"].setdefault("processing_strategy", report.source_type)
        payload["sections"] = payload.get("sections", build_section_status(charts_ready=bool(payload.get("charts"))))
        return sanitize_for_json(payload)

    base_payload = copy.deepcopy(cache_entry.full_payload or cache_entry.preview_payload or report.report_payload or {})
    base_payload.setdefault("overview", {})
    base_payload.setdefault("metadata", {})
    base_payload["dataset_name"] = report.dataset_name
    base_payload["source_type"] = report.source_type
    base_payload["target_column"] = report.target_column
    base_payload["charts"] = base_payload.get("charts", [])
    base_payload["metadata"].update(
        {
            "is_preview": cache_entry.status != "completed" or cache_entry.full_payload is None,
            "processing_mode": cache_entry.processing_mode,
            "file_type": cache_entry.file_type,
            "file_size_bytes": cache_entry.file_size_bytes,
            "cache_hit": from_cache,
            "job_id": cache_entry.id,
            "optimized_mode": is_optimized_mode(cache_entry.processing_mode),
            "processing_strategy": build_processing_strategy(cache_entry.processing_mode),
            "max_upload_size_bytes": settings.max_upload_size_bytes,
            "storage_backend": cache_entry.storage_backend,
        }
    )
    base_payload["sections"] = build_materialized_section_status(cache_entry, base_payload)
    return sanitize_for_json(base_payload)


def sync_report_from_cache(
    report: AnalysisReport,
    cache_entry: AnalysisCacheEntry,
    *,
    from_cache: bool = False,
) -> None:
    payload = materialize_report_payload(report, cache_entry, from_cache=from_cache)
    overview = payload.get("overview", {})
    report.status = cache_entry.status
    report.row_count = int(cache_entry.row_count or overview.get("row_count") or report.row_count or 0)
    report.column_count = int(cache_entry.column_count or overview.get("column_count") or report.column_count or 0)
    report.report_payload = payload
    report.notes = cache_entry.error_message or cache_entry.progress_message


def build_materialized_section_status(
    cache_entry: AnalysisCacheEntry,
    payload: dict[str, Any],
) -> dict[str, bool]:
    base_status = build_section_status(charts_ready=bool(payload.get("charts")))
    base_status.update(payload.get("sections") or {})
    base_status.update(cache_entry.sections_ready or {})
    base_status["rows"] = True
    return base_status


def propagate_cache_state_to_reports(db: Session, cache_entry: AnalysisCacheEntry) -> None:
    refreshed_entry = db.scalar(
        select(AnalysisCacheEntry)
        .where(AnalysisCacheEntry.id == cache_entry.id)
        .options(
            selectinload(AnalysisCacheEntry.report_links).selectinload(AnalysisReportCacheLink.report),
        )
    )
    if not refreshed_entry:
        return

    for link in refreshed_entry.report_links:
        sync_report_from_cache(link.report, refreshed_entry)


def build_preview_payload(
    *,
    stored_upload: StoredUpload,
    dataset_name: str,
    target_column: str | None,
) -> dict[str, Any]:
    preview_frame = load_preview_sample(
        source_path=stored_upload.storage_path,
        file_type=stored_upload.file_type,
        extension=stored_upload.extension,
        limit=settings.preview_sample_rows,
        sample=stored_upload.processing_mode == "large",
    )
    if preview_frame.empty:
        raise ValueError("The dataset is empty after parsing. Please upload a file with records.")

    payload = analyze_dataframe(
        preview_frame,
        dataset_name=dataset_name,
        source_type="upload",
        target_column=target_column,
        include_charts=False,
        metadata={
            "is_preview": stored_upload.processing_mode != "small",
            "processing_mode": stored_upload.processing_mode,
            "file_type": stored_upload.file_type,
            "file_size_bytes": stored_upload.file_size_bytes,
            "sample_row_count": int(len(preview_frame)),
            "optimized_mode": is_optimized_mode(stored_upload.processing_mode),
            "processing_strategy": build_processing_strategy(stored_upload.processing_mode),
            "sample_strategy": "reservoir" if stored_upload.processing_mode == "large" else "head",
            "max_upload_size_bytes": settings.max_upload_size_bytes,
        },
    )
    payload["charts"] = []
    payload["sections"] = build_section_status(charts_ready=False)
    return sanitize_for_json(payload)


def build_small_file_payload(
    *,
    stored_upload: StoredUpload,
    dataset_name: str,
    target_column: str | None,
) -> dict[str, Any]:
    with stored_upload.storage_path.open("rb") as upload_stream:
        dataframe = parse_uploaded_dataframe(stored_upload.original_filename, upload_stream)

    payload = analyze_dataframe(
        dataframe,
        dataset_name=dataset_name,
        source_type="upload",
        target_column=target_column,
        include_charts=False,
        metadata={
            "is_preview": False,
            "processing_mode": stored_upload.processing_mode,
            "file_type": stored_upload.file_type,
            "file_size_bytes": stored_upload.file_size_bytes,
            "sample_row_count": int(len(dataframe)),
            "optimized_mode": is_optimized_mode(stored_upload.processing_mode),
            "processing_strategy": build_processing_strategy(stored_upload.processing_mode),
            "sample_strategy": "full_dataframe",
            "max_upload_size_bytes": settings.max_upload_size_bytes,
        },
    )
    payload["charts"] = []
    payload["sections"] = build_section_status(charts_ready=False)
    return sanitize_for_json(payload)


def build_large_sample_payload(
    *,
    sample_frame: pd.DataFrame,
    dataset_name: str,
    target_column: str | None,
    processing_mode: str,
    file_type: str,
    file_size_bytes: int | None,
) -> dict[str, Any]:
    if sample_frame.empty:
        raise ValueError("The uploaded dataset did not contain any readable rows.")

    cleaned_frame, cleaning_summary = clean_dataframe(sample_frame.copy())
    if cleaned_frame.empty:
        raise ValueError("The dataset is empty after cleaning. Please provide rows with actual values.")

    normalized_target = normalize_column_name(target_column) if target_column else None
    if normalized_target and normalized_target not in cleaned_frame.columns:
        raise ValueError(
            f"Target column '{target_column}' was not found after cleaning. "
            "Use one of the cleaned column names shown in the dashboard."
        )

    overview = build_overview(
        sample_frame,
        cleaned_frame,
        cleaning_summary,
        normalized_target,
    )

    return {
        "dataset_name": dataset_name,
        "source_type": "upload",
        "target_column": normalized_target,
        "overview": overview,
        "cleaning": cleaning_summary,
        "trends": analyze_trends(cleaned_frame),
        "modeling": build_modeling_summary(cleaned_frame, normalized_target),
        "metadata": {
            "is_preview": False,
            "processing_mode": processing_mode,
            "file_type": file_type,
            "file_size_bytes": file_size_bytes,
            "sample_row_count": int(len(cleaned_frame)),
        },
    }


def fetch_single_row_mapping(
    connection: duckdb.DuckDBPyConnection,
    query: str,
) -> dict[str, Any]:
    result = connection.execute(query)
    row = result.fetchone()
    if row is None:
        return {}

    return {
        str(column_metadata[0]): value
        for column_metadata, value in zip(result.description or [], row)
    }


def process_cache_entry(cache_entry_id: str) -> None:
    try:
        logger.info(f"Processing cache entry {cache_entry_id} started")
        with SessionLocal() as db:
            cache_entry = db.get(AnalysisCacheEntry, cache_entry_id)
            if settings.analytics_memory_limit_mb:
                import resource
                resource.setrlimit(resource.RLIMIT_AS, (settings.analytics_memory_limit_mb * 1024 * 1024, -1))
            if not cache_entry:
                logger.warning("Cache entry not found for background job: cache_entry_id=%s", cache_entry_id)
                return
            if cache_entry.status == "completed" and cache_entry.full_payload:
                logger.info("Background job skipped for completed cache entry: cache_entry_id=%s", cache_entry_id)
                return

            logger.info(
                "Background analytics started: cache_entry_id=%s mode=%s type=%s size_bytes=%s",
                cache_entry.id,
                cache_entry.processing_mode,
                cache_entry.file_type,
                cache_entry.file_size_bytes,
            )
            cache_entry.status = "processing"
            cache_entry.progress = max(cache_entry.progress, 38)
            cache_entry.progress_message = (
                "Converting chunked data for full analytics."
                if cache_entry.processing_mode == "medium"
                else "Large dataset detected. Processing in optimized mode."
            )
            cache_entry.error_message = None
            cache_entry.started_at = cache_entry.started_at or utcnow()
            cache_entry.completed_at = None
            cache_entry.failed_at = None
            db.commit()

            if should_convert_to_parquet(cache_entry):
                parquet_path = convert_to_parquet_if_needed(cache_entry)
                if parquet_path:
                    cache_entry.parquet_path = str(parquet_path)
                    cache_entry.progress = 56
                    cache_entry.progress_message = "Optimized Parquet storage is ready."
                    db.commit()
                    logger.info(
                        "Parquet conversion complete: cache_entry_id=%s parquet_path=%s",
                        cache_entry.id,
                        parquet_path,
                    )

            full_payload = build_full_payload(cache_entry)
            cache_entry.full_payload = full_payload
            logger.info(f"Full payload built for {cache_entry_id}")
            cache_entry.row_count = int(full_payload.get("overview", {}).get("row_count", cache_entry.row_count))
            cache_entry.column_count = int(
                full_payload.get("overview", {}).get("column_count", cache_entry.column_count)
            )
            cache_entry.sections_ready = full_payload.get("sections", build_section_status(charts_ready=False))
            cache_entry.status = "completed"
            cache_entry.progress = 100
            cache_entry.progress_message = "Full analytics complete."
            cache_entry.error_message = None
            cache_entry.completed_at = utcnow()
            propagate_cache_state_to_reports(db, cache_entry)
            db.commit()
            logger.info(
                "Background analytics completed: cache_entry_id=%s rows=%s columns=%s",
                cache_entry.id,
                cache_entry.row_count,
                cache_entry.column_count,
            )
    except MemoryError as mem_exc:
        logger.error("MemoryError in processing %s: %s", cache_entry_id, mem_exc)
        raise
    except Exception as exc:
        logger.exception("Background analytics failed: cache_entry_id=%s", cache_entry_id)
        with SessionLocal() as db:
            cache_entry = db.get(AnalysisCacheEntry, cache_entry_id)
            if not cache_entry:
                return
            cache_entry.status = "failed"
            cache_entry.progress = 100
            cache_entry.error_message = f"Processing error: {str(exc)[:200]}"
            cache_entry.progress_message = "Analytics failed due to processing error. Try smaller sample."
            cache_entry.failed_at = utcnow()
            propagate_cache_state_to_reports(db, cache_entry)
            db.commit()
        raise  
def build_full_payload(cache_entry: AnalysisCacheEntry) -> dict[str, Any]:
    canonical_dataset_name = Path(cache_entry.original_filename).stem or "dataset"
    with managed_dataset_connection(cache_entry) as connection:
        sample_frame = load_sample_frame_from_connection(
            connection,
            settings.analytics_sample_rows,
        )
        sample_payload = build_large_sample_payload(
            sample_frame=sample_frame,
            dataset_name=canonical_dataset_name,
            target_column=cache_entry.target_column,
            processing_mode=cache_entry.processing_mode,
            file_type=cache_entry.file_type,
            file_size_bytes=cache_entry.file_size_bytes,
        )
        column_definitions = build_column_definitions(sample_frame, sample_payload)
        row_count = int(connection.execute("SELECT COUNT(*) FROM dataset_source").fetchone()[0])
        column_profiles = build_column_profiles(connection, row_count, column_definitions, sample_payload)
        all_null_columns = {profile["column"] for profile in column_profiles if profile["missing_count"] >= row_count}
        active_columns = [profile for profile in column_profiles if profile["column"] not in all_null_columns]
        active_definitions = [definition for definition in column_definitions if definition.normalized_name not in all_null_columns]

        summary_statistics = build_large_summary_statistics(
            connection=connection,
            row_count=row_count,
            column_definitions=active_definitions,
            column_profiles=active_columns,
        )
        correlations = build_large_correlations(connection, active_definitions)
        outliers = build_large_outliers(connection, active_definitions)
        trends = sample_payload.get("trends", [])
        modeling = sample_payload.get("modeling", {})
        detected_data_types = {
            definition.normalized_name: definition.display_dtype for definition in active_definitions
        }
        cleaning = {
            "original_shape": {"rows": row_count, "columns": len(column_definitions)},
            "cleaned_shape": {"rows": row_count, "columns": len(active_definitions)},
            "column_mapping": {
                definition.original_name: definition.normalized_name for definition in column_definitions
            },
            "removed_all_null_columns": sorted(all_null_columns),
            "empty_rows_dropped": 0,
            "duplicate_rows_removed": 0,
            "missing_values_before": int(sum(profile["missing_count"] for profile in active_columns)),
            "missing_values_after": int(sum(profile["missing_count"] for profile in active_columns)),
            "detected_data_types": detected_data_types,
        }
        overview = {
            "row_count": row_count,
            "column_count": len(active_definitions),
            "original_row_count": row_count,
            "original_column_count": len(column_definitions),
            "target_column": cache_entry.target_column,
            "preview_rows": sample_payload.get("overview", {}).get("preview_rows", []),
            "columns": active_columns,
            "detected_data_types": detected_data_types,
        }
        insights, recommendations = build_narrative(
            overview=overview,
            cleaning=cleaning,
            correlations=correlations,
            outliers=outliers,
            trends=trends,
            modeling=modeling,
        )

        payload = {
            "dataset_name": canonical_dataset_name,
            "source_type": "upload",
            "target_column": cache_entry.target_column,
            "overview": overview,
            "cleaning": cleaning,
            "summary_statistics": summary_statistics,
            "correlations": correlations,
            "outliers": outliers,
            "trends": trends,
            "charts": [],
            "modeling": modeling,
            "insights": insights,
            "recommendations": recommendations,
            "metadata": {
                "is_preview": False,
                "processing_mode": cache_entry.processing_mode,
                "file_type": cache_entry.file_type,
                "file_size_bytes": cache_entry.file_size_bytes,
                "sample_row_count": int(len(sample_frame)),
                "optimized_mode": is_optimized_mode(cache_entry.processing_mode),
                "processing_strategy": build_processing_strategy(cache_entry.processing_mode),
                "sample_strategy": "reservoir",
                "max_upload_size_bytes": settings.max_upload_size_bytes,
            },
            "sections": build_section_status(charts_ready=False),
        }
        return sanitize_for_json(payload)


def build_column_definitions(
    sample_frame: pd.DataFrame,
    sample_payload: dict[str, Any],
) -> list[ColumnDefinition]:
    column_mapping = sample_payload.get("cleaning", {}).get("column_mapping", {})
    detected_data_types = sample_payload.get("cleaning", {}).get("detected_data_types", {})
    profile_lookup = {
        profile["column"]: profile for profile in sample_payload.get("overview", {}).get("columns", [])
    }

    definitions: list[ColumnDefinition] = []
    for original_name in [str(column) for column in sample_frame.columns]:
        normalized_name = column_mapping.get(original_name, normalize_column_name(original_name))
        display_dtype = detected_data_types.get(
            normalized_name,
            profile_lookup.get(normalized_name, {}).get("dtype", "object"),
        )
        inferred_kind = infer_column_kind(display_dtype)
        definitions.append(
            ColumnDefinition(
                original_name=original_name,
                normalized_name=normalized_name,
                inferred_kind=inferred_kind,
                display_dtype=display_dtype,
            )
        )

    return definitions


def infer_column_kind(display_dtype: str) -> str:
    normalized = display_dtype.lower()
    if "datetime" in normalized or "timestamp" in normalized or normalized.startswith("date"):
        return "datetime"
    if any(token in normalized for token in ("int", "float", "double", "decimal")):
        return "numeric"
    return "categorical"


def build_column_profiles(
    connection: duckdb.DuckDBPyConnection,
    row_count: int,
    column_definitions: list[ColumnDefinition],
    sample_payload: dict[str, Any],
) -> list[dict[str, Any]]:
    if not column_definitions:
        return []

    sample_profiles = {
        profile["column"]: profile for profile in sample_payload.get("overview", {}).get("columns", [])
    }
    aggregate_expressions: list[str] = []
    for index, definition in enumerate(column_definitions):
        identifier = quote_identifier(definition.original_name)
        aggregate_expressions.extend(
            [
                (
                    f"SUM(CASE WHEN {missing_condition_sql(definition.original_name)} THEN 1 ELSE 0 END) "
                    f"AS missing_count__{index}"
                ),
                (
                    "COUNT(DISTINCT CASE WHEN "
                    f"{missing_condition_sql(definition.original_name)} THEN NULL "
                    f"ELSE {identifier} END) AS unique_values__{index}"
                ),
            ]
        )

    aggregate_values = fetch_single_row_mapping(
        connection,
        f"SELECT {', '.join(aggregate_expressions)} FROM dataset_source",
    )
    profiles: list[dict[str, Any]] = []

    for index, definition in enumerate(column_definitions):
        missing_count = int(aggregate_values.get(f"missing_count__{index}") or 0)
        unique_values = int(aggregate_values.get(f"unique_values__{index}") or 0)
        sample_values = sample_profiles.get(definition.normalized_name, {}).get("sample_values", [])
        profiles.append(
            {
                "column": definition.normalized_name,
                "dtype": definition.display_dtype,
                "missing_count": missing_count,
                "missing_percentage": round((missing_count / row_count * 100), 2) if row_count else 0.0,
                "unique_values": unique_values,
                "sample_values": sample_values,
            }
        )

    return profiles


def build_large_numeric_summary_lookup(
    connection: duckdb.DuckDBPyConnection,
    column_definitions: list[ColumnDefinition],
) -> dict[str, dict[str, Any]]:
    numeric_definitions = [definition for definition in column_definitions if definition.inferred_kind == "numeric"]
    if not numeric_definitions:
        return {}

    typed_columns = [
        f"TRY_CAST({quote_identifier(definition.original_name)} AS DOUBLE) AS value__{index}"
        for index, definition in enumerate(numeric_definitions)
    ]
    aggregate_expressions: list[str] = []
    for index in range(len(numeric_definitions)):
        aggregate_expressions.extend(
            [
                f"AVG(value__{index}) AS mean__{index}",
                f"QUANTILE_CONT(value__{index}, 0.5) AS median__{index}",
                f"STDDEV_POP(value__{index}) AS std__{index}",
                f"MIN(value__{index}) AS min__{index}",
                f"MAX(value__{index}) AS max__{index}",
                f"QUANTILE_CONT(value__{index}, 0.25) AS q1__{index}",
                f"QUANTILE_CONT(value__{index}, 0.75) AS q3__{index}",
            ]
        )

    aggregate_values = fetch_single_row_mapping(
        connection,
        (
            "WITH typed AS ("
            f"SELECT {', '.join(typed_columns)} FROM dataset_source"
            ") "
            f"SELECT {', '.join(aggregate_expressions)} FROM typed"
        ),
    )

    lookup: dict[str, dict[str, Any]] = {}
    for index, definition in enumerate(numeric_definitions):
        min_value = aggregate_values.get(f"min__{index}")
        if min_value is None:
            continue

        lookup[definition.normalized_name] = {
            "mean": round(float(aggregate_values.get(f"mean__{index}")), 4)
            if aggregate_values.get(f"mean__{index}") is not None
            else None,
            "median": round(float(aggregate_values.get(f"median__{index}")), 4)
            if aggregate_values.get(f"median__{index}") is not None
            else None,
            "std": round(float(aggregate_values.get(f"std__{index}")), 4)
            if aggregate_values.get(f"std__{index}") is not None
            else None,
            "min": round(float(min_value), 4),
            "max": round(float(aggregate_values.get(f"max__{index}")), 4)
            if aggregate_values.get(f"max__{index}") is not None
            else None,
            "q1": round(float(aggregate_values.get(f"q1__{index}")), 4)
            if aggregate_values.get(f"q1__{index}") is not None
            else None,
            "q3": round(float(aggregate_values.get(f"q3__{index}")), 4)
            if aggregate_values.get(f"q3__{index}") is not None
            else None,
        }

    return lookup


def build_large_datetime_summary_lookup(
    connection: duckdb.DuckDBPyConnection,
    column_definitions: list[ColumnDefinition],
) -> dict[str, dict[str, Any]]:
    datetime_definitions = [definition for definition in column_definitions if definition.inferred_kind == "datetime"]
    if not datetime_definitions:
        return {}

    typed_columns = [
        f"TRY_CAST({quote_identifier(definition.original_name)} AS TIMESTAMP) AS value__{index}"
        for index, definition in enumerate(datetime_definitions)
    ]
    aggregate_expressions: list[str] = []
    for index in range(len(datetime_definitions)):
        aggregate_expressions.extend(
            [
                f"MIN(value__{index}) AS min__{index}",
                f"MAX(value__{index}) AS max__{index}",
            ]
        )

    aggregate_values = fetch_single_row_mapping(
        connection,
        (
            "WITH typed AS ("
            f"SELECT {', '.join(typed_columns)} FROM dataset_source"
            ") "
            f"SELECT {', '.join(aggregate_expressions)} FROM typed"
        ),
    )

    lookup: dict[str, dict[str, Any]] = {}
    for index, definition in enumerate(datetime_definitions):
        lookup[definition.normalized_name] = {
            "min": aggregate_values.get(f"min__{index}").isoformat()
            if aggregate_values.get(f"min__{index}") is not None
            else None,
            "max": aggregate_values.get(f"max__{index}").isoformat()
            if aggregate_values.get(f"max__{index}") is not None
            else None,
        }

    return lookup


def build_large_summary_statistics(
    *,
    connection: duckdb.DuckDBPyConnection,
    row_count: int,
    column_definitions: list[ColumnDefinition],
    column_profiles: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    profile_lookup = {profile["column"]: profile for profile in column_profiles}
    numeric_summary_lookup = build_large_numeric_summary_lookup(connection, column_definitions)
    datetime_summary_lookup = build_large_datetime_summary_lookup(connection, column_definitions)
    summary_statistics: list[dict[str, Any]] = []

    for definition in column_definitions:
        profile = profile_lookup[definition.normalized_name]
        base_stats: dict[str, Any] = {
            "column": definition.normalized_name,
            "dtype": definition.display_dtype,
            "non_null_count": int(max(row_count - profile["missing_count"], 0)),
            "unique_values": profile["unique_values"],
        }
        identifier = quote_identifier(definition.original_name)

        if definition.inferred_kind == "numeric":
            base_stats.update(numeric_summary_lookup.get(definition.normalized_name, {}))
        elif definition.inferred_kind == "datetime":
            base_stats.update(datetime_summary_lookup.get(definition.normalized_name, {}))
        else:
            top_value_query = f"""
                SELECT CAST({identifier} AS VARCHAR) AS value, COUNT(*) AS frequency
                FROM dataset_source
                WHERE NOT ({missing_condition_sql(definition.original_name)})
                GROUP BY 1
                ORDER BY frequency DESC, value ASC
                LIMIT 1
            """
            top_row = connection.execute(top_value_query).fetchone()
            base_stats.update(
                {
                    "top_value": top_row[0] if top_row else None,
                    "top_frequency": int(top_row[1]) if top_row else 0,
                }
            )

        summary_statistics.append(sanitize_for_json(base_stats))

    return summary_statistics


def build_large_correlations(
    connection: duckdb.DuckDBPyConnection,
    column_definitions: list[ColumnDefinition],
) -> dict[str, Any]:
    numeric_columns = [definition for definition in column_definitions if definition.inferred_kind == "numeric"][
        :MAX_CORRELATION_COLUMNS
    ]
    if len(numeric_columns) < 2:
        return {"available": False, "columns": [], "matrix": [], "strongest_pairs": []}

    correlation_expressions: list[str] = []
    for left_index, left_definition in enumerate(numeric_columns):
        left_identifier = quote_identifier(left_definition.original_name)
        for right_index, right_definition in enumerate(numeric_columns[left_index + 1 :], start=left_index + 1):
            right_identifier = quote_identifier(right_definition.original_name)
            correlation_expressions.append(
                f"CORR(TRY_CAST({left_identifier} AS DOUBLE), TRY_CAST({right_identifier} AS DOUBLE)) AS corr_{left_index}_{right_index}"
            )

    correlation_row = connection.execute(
        f"SELECT {', '.join(correlation_expressions)} FROM dataset_source"
    ).fetchone()
    correlation_values = list(correlation_row) if correlation_row else []
    matrix = [
        [1.0 if row_index == column_index else 0.0 for column_index in range(len(numeric_columns))]
        for row_index in range(len(numeric_columns))
    ]
    strongest_pairs: list[dict[str, Any]] = []

    value_index = 0
    for left_index, left_definition in enumerate(numeric_columns):
        for right_index, right_definition in enumerate(numeric_columns[left_index + 1 :], start=left_index + 1):
            raw_value = correlation_values[value_index] if value_index < len(correlation_values) else None
            value_index += 1
            correlation_value = round(float(raw_value), 3) if raw_value is not None else 0.0
            matrix[left_index][right_index] = correlation_value
            matrix[right_index][left_index] = correlation_value
            strongest_pairs.append(
                {
                    "left_column": left_definition.normalized_name,
                    "right_column": right_definition.normalized_name,
                    "correlation": correlation_value,
                }
            )

    strongest_pairs = sorted(strongest_pairs, key=lambda pair: abs(pair["correlation"]), reverse=True)[:6]
    return {
        "available": True,
        "columns": [definition.normalized_name for definition in numeric_columns],
        "matrix": matrix,
        "strongest_pairs": strongest_pairs,
    }


def build_large_outliers(
    connection: duckdb.DuckDBPyConnection,
    column_definitions: list[ColumnDefinition],
) -> list[dict[str, Any]]:
    numeric_definitions = [item for item in column_definitions if item.inferred_kind == "numeric"]
    if not numeric_definitions:
        return []

    typed_columns = [
        f"TRY_CAST({quote_identifier(definition.original_name)} AS DOUBLE) AS value__{index}"
        for index, definition in enumerate(numeric_definitions)
    ]
    quartile_expressions: list[str] = []
    for index in range(len(numeric_definitions)):
        quartile_expressions.extend(
            [
                f"QUANTILE_CONT(value__{index}, 0.25) AS q1__{index}",
                f"QUANTILE_CONT(value__{index}, 0.75) AS q3__{index}",
                f"COUNT(value__{index}) AS non_null_count__{index}",
            ]
        )

    quartile_values = fetch_single_row_mapping(
        connection,
        (
            "WITH typed AS ("
            f"SELECT {', '.join(typed_columns)} FROM dataset_source"
            ") "
            f"SELECT {', '.join(quartile_expressions)} FROM typed"
        ),
    )

    bounds_by_index: dict[int, dict[str, float | int | str]] = {}
    for index, definition in enumerate(numeric_definitions):
        q1_value = quartile_values.get(f"q1__{index}")
        q3_value = quartile_values.get(f"q3__{index}")
        non_null_count = int(quartile_values.get(f"non_null_count__{index}") or 0)
        if q1_value is None or q3_value is None or non_null_count < 4:
            continue

        iqr = float(q3_value) - float(q1_value)
        if iqr == 0:
            continue

        bounds_by_index[index] = {
            "column": definition.normalized_name,
            "non_null_count": non_null_count,
            "lower_bound": float(q1_value) - (1.5 * iqr),
            "upper_bound": float(q3_value) + (1.5 * iqr),
        }

    if not bounds_by_index:
        return []

    outlier_count_expressions = [
        (
            "SUM(CASE WHEN value__{index} IS NOT NULL AND "
            "(value__{index} < {lower_bound} OR value__{index} > {upper_bound}) "
            "THEN 1 ELSE 0 END) AS outlier_count__{index}"
        ).format(
            index=index,
            lower_bound=bounds["lower_bound"],
            upper_bound=bounds["upper_bound"],
        )
        for index, bounds in bounds_by_index.items()
    ]
    outlier_count_values = fetch_single_row_mapping(
        connection,
        (
            "WITH typed AS ("
            f"SELECT {', '.join(typed_columns)} FROM dataset_source"
            ") "
            f"SELECT {', '.join(outlier_count_expressions)} FROM typed"
        ),
    )

    outliers: list[dict[str, Any]] = []
    for index, bounds in bounds_by_index.items():
        outlier_count = int(outlier_count_values.get(f"outlier_count__{index}") or 0)
        if not outlier_count:
            continue

        non_null_count = int(bounds["non_null_count"])
        lower_bound = float(bounds["lower_bound"])
        upper_bound = float(bounds["upper_bound"])
        outliers.append(
            {
                "column": bounds["column"],
                "count": outlier_count,
                "percentage": round(outlier_count / max(non_null_count, 1) * 100, 2),
                "lower_bound": round(lower_bound, 4),
                "upper_bound": round(upper_bound, 4),
            }
        )

    return outliers


def ensure_report_section(
    db: Session,
    report: AnalysisReport,
    section_name: str,
) -> Any:
    if section_name not in SUPPORTED_REPORT_SECTIONS:
        raise ValueError("Requested report section is not supported.")

    cache_entry = get_cache_entry_for_report(db, report.id)
    if not cache_entry:
        return (report.report_payload or {}).get(section_name)

    if section_name != "charts":
        payload = materialize_report_payload(report, cache_entry)
        return payload.get(section_name)

    if cache_entry.full_payload and cache_entry.full_payload.get("charts"):
        return cache_entry.full_payload["charts"]

    if cache_entry.status == "completed" and cache_entry.full_payload is not None:
        charts = generate_charts_for_cache_entry(cache_entry)
        cache_entry.full_payload["charts"] = charts
        cache_entry.sections_ready = {
            **(cache_entry.sections_ready or {}),
            "charts": True,
        }
        propagate_cache_state_to_reports(db, cache_entry)
        db.commit()
        return charts

    preview_payload = materialize_report_payload(report, cache_entry)
    if preview_payload.get("charts"):
        return preview_payload["charts"]

    return generate_transient_preview_charts(cache_entry)


def get_report_rows_page(
    db: Session,
    report: AnalysisReport,
    *,
    page: int,
    page_size: int,
) -> dict[str, Any]:
    normalized_page = max(page, 1)
    normalized_page_size = min(max(page_size, 1), settings.max_table_page_size)
    cache_entry = get_cache_entry_for_report(db, report.id)

    if not cache_entry:
        payload = report.report_payload or {}
        preview_rows = payload.get("overview", {}).get("preview_rows", [])
        columns = [column.get("column") for column in payload.get("overview", {}).get("columns", [])]
        total_rows = len(preview_rows)
        start_index = (normalized_page - 1) * normalized_page_size
        end_index = start_index + normalized_page_size
        return {
            "page": normalized_page,
            "page_size": normalized_page_size,
            "total_rows": total_rows,
            "total_pages": max((total_rows + normalized_page_size - 1) // normalized_page_size, 1),
            "columns": columns,
            "rows": preview_rows[start_index:end_index],
            "is_preview": True,
        }

    if cache_entry.file_type == "excel" and not cache_entry.parquet_path:
        payload = materialize_report_payload(report, cache_entry)
        preview_rows = payload.get("overview", {}).get("preview_rows", [])
        columns = [column.get("column") for column in payload.get("overview", {}).get("columns", [])]
        total_rows = len(preview_rows)
        start_index = (normalized_page - 1) * normalized_page_size
        end_index = start_index + normalized_page_size
        return {
            "page": normalized_page,
            "page_size": normalized_page_size,
            "total_rows": total_rows,
            "total_pages": max((total_rows + normalized_page_size - 1) // normalized_page_size, 1),
            "columns": columns,
            "rows": preview_rows[start_index:end_index],
            "is_preview": True,
        }

    payload = materialize_report_payload(report, cache_entry)
    columns = [column.get("column") for column in payload.get("overview", {}).get("columns", [])]
    if not columns:
        return {
            "page": normalized_page,
            "page_size": normalized_page_size,
            "total_rows": 0,
            "total_pages": 1,
            "columns": [],
            "rows": [],
            "is_preview": True,
        }

    with managed_dataset_connection(cache_entry) as connection:
        offset = (normalized_page - 1) * normalized_page_size
        column_definitions = build_column_definitions_from_payload(payload, connection)
        select_list = ", ".join(
            f"{quote_identifier(definition.original_name)} AS {quote_identifier(definition.normalized_name)}"
            for definition in column_definitions
        )
        rows = connection.execute(
            f"SELECT {select_list} FROM dataset_source LIMIT {normalized_page_size} OFFSET {offset}"
        ).fetchdf()
        total_rows = cache_entry.row_count or int(connection.execute("SELECT COUNT(*) FROM dataset_source").fetchone()[0])
        return {
            "page": normalized_page,
            "page_size": normalized_page_size,
            "total_rows": total_rows,
            "total_pages": max((total_rows + normalized_page_size - 1) // normalized_page_size, 1),
            "columns": [definition.normalized_name for definition in column_definitions],
            "rows": sanitize_for_json(rows.to_dict(orient="records")),
            "is_preview": cache_entry.status != "completed",
        }
def build_column_definitions_from_payload(
    payload: dict[str, Any],
    connection: duckdb.DuckDBPyConnection,
) -> list[ColumnDefinition]:
    schema_rows = connection.execute("DESCRIBE SELECT * FROM dataset_source").fetchall()
    mapping = payload.get("cleaning", {}).get("column_mapping", {})
    reverse_mapping = {original: normalized for original, normalized in mapping.items()}
    detected_data_types = payload.get("cleaning", {}).get("detected_data_types", {})
    definitions: list[ColumnDefinition] = []
    for row in schema_rows:
        original_name = str(row[0])
        normalized_name = reverse_mapping.get(original_name, normalize_column_name(original_name))
        display_dtype = detected_data_types.get(normalized_name, str(row[1]))
        definitions.append(
            ColumnDefinition(
                original_name=original_name,
                normalized_name=normalized_name,
                inferred_kind=infer_column_kind(display_dtype),
                display_dtype=display_dtype,
            )
        )
    return definitions
def get_cache_entry_for_report(db: Session, report_id: str) -> AnalysisCacheEntry | None:
    statement = (
        select(AnalysisCacheEntry)
        .join(AnalysisReportCacheLink, AnalysisReportCacheLink.cache_entry_id == AnalysisCacheEntry.id)
        .where(AnalysisReportCacheLink.report_id == report_id)
    )
    return db.scalar(statement)
def should_convert_to_parquet(cache_entry: AnalysisCacheEntry) -> bool:
    if cache_entry.file_type == "csv":
        return cache_entry.processing_mode in {"medium", "large"}
    return True
def convert_to_parquet_if_needed(cache_entry: AnalysisCacheEntry) -> Path | None:
    destination = build_parquet_path(cache_entry.content_hash)
    if destination.exists():
        return destination

    source_path = Path(cache_entry.storage_path)
    source_path = resolve_cache_entry_source_path(cache_entry)
    destination.parent.mkdir(parents=True, exist_ok=True)

    if cache_entry.file_type == "csv":
        convert_relation_file_to_parquet(source_path, destination, "csv")
        return destination
    if cache_entry.file_type == "json":
        convert_relation_file_to_parquet(source_path, destination, "json")
        return destination
    if cache_entry.file_type == "excel":
        extension = source_path.suffix.lower()
        if extension == ".xls":
            convert_legacy_excel_to_parquet(source_path, destination)
        else:
            convert_modern_excel_to_parquet(source_path, destination)
        return destination

    return None
def convert_relation_file_to_parquet(source_path: Path, destination: Path, file_type: str) -> None:
    relation_sql = build_relation_sql_from_path(source_path, file_type=file_type)
    connection = duckdb.connect()
    try:
        connection.execute(
            f"COPY (SELECT * FROM {relation_sql}) TO {sql_literal(str(destination))} (FORMAT PARQUET, COMPRESSION ZSTD)"
        )
    finally:
        connection.close()
def convert_modern_excel_to_parquet(source_path: Path, destination: Path) -> None:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    writer: pq.ParquetWriter | None = None
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            raise ValueError("The uploaded Excel file is empty.")

        columns = build_excel_columns(header_row)
        chunk_rows: list[list[Any]] = []
        for row in rows:
            padded_row = list(row[: len(columns)])
            if len(padded_row) < len(columns):
                padded_row.extend([None] * (len(columns) - len(padded_row)))
            chunk_rows.append(padded_row)
            if len(chunk_rows) >= settings.excel_chunk_rows:
                writer = write_excel_chunk(writer, destination, columns, chunk_rows)
                chunk_rows = []

        if chunk_rows:
            writer = write_excel_chunk(writer, destination, columns, chunk_rows)

        if writer is None:
            writer = pq.ParquetWriter(destination, pa.schema([(column, pa.string()) for column in columns]))
            empty_arrays = [pa.array([], type=pa.string()) for _ in columns]
            writer.write_table(pa.Table.from_arrays(empty_arrays, names=columns))
    finally:
        workbook.close()
        if writer is not None:
            writer.close()
def convert_legacy_excel_to_parquet(source_path: Path, destination: Path) -> None:
    writer: pq.ParquetWriter | None = None
    rows_to_skip = 0
    columns: list[str] | None = None

    try:
        while True:
            chunk = pd.read_excel(
                source_path,
                nrows=settings.excel_chunk_rows,
                skiprows=(lambda row_index, offset=rows_to_skip: 0 < row_index <= offset) if rows_to_skip else None,
            )
            if chunk.empty:
                break

            if columns is None:
                columns = [str(column) for column in chunk.columns]

            chunk = chunk.reindex(columns=columns)
            writer = write_dataframe_chunk(writer, destination, columns, chunk)
            rows_to_skip += len(chunk)
            if len(chunk) < settings.excel_chunk_rows:
                break

        if writer is None:
            if columns is None:
                raise ValueError("The uploaded Excel file is empty.")
            writer = pq.ParquetWriter(destination, pa.schema([(column, pa.string()) for column in columns]))
            empty_arrays = [pa.array([], type=pa.string()) for _ in columns]
            writer.write_table(pa.Table.from_arrays(empty_arrays, names=columns))
    finally:
        if writer is not None:
            writer.close()
def write_excel_chunk(
    writer: pq.ParquetWriter | None,
    destination: Path,
    columns: list[str],
    chunk_rows: list[list[Any]],
) -> pq.ParquetWriter:
    data = {column: [] for column in columns}
    for row in chunk_rows:
        for column, value in zip(columns, row):
            data[column].append(stringify_cell_value(value))
    table = pa.Table.from_pydict(
        {column: pa.array(values, type=pa.string()) for column, values in data.items()}
    )
    if writer is None:
        writer = pq.ParquetWriter(destination, table.schema)
    writer.write_table(table)
    return writer
def write_dataframe_chunk(
    writer: pq.ParquetWriter | None,
    destination: Path,
    columns: list[str],
    chunk: pd.DataFrame,
) -> pq.ParquetWriter:
    data = {}
    for column in columns:
        values = [stringify_cell_value(value) for value in chunk[column].tolist()]
        data[column] = pa.array(values, type=pa.string())
    table = pa.Table.from_pydict(data)
    if writer is None:
        writer = pq.ParquetWriter(destination, table.schema)
    writer.write_table(table)
    return writer
def stringify_cell_value(value: Any) -> str | None:
    if value is None:
        return None
    if pd.isna(value):
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except TypeError:
            pass
    text = str(value).strip()
    return text or None


def to_sql_number(value: float) -> str:
    return f"{float(value):.15g}"


def format_chart_tick(value: float) -> str:
    return f"{float(value):.4g}"


def build_exact_categorical_charts(
    connection: duckdb.DuckDBPyConnection,
    definition: ColumnDefinition,
) -> list[dict[str, Any]]:
    identifier = quote_identifier(definition.original_name)
    counts_query = f"""
        SELECT CAST({identifier} AS VARCHAR) AS category, COUNT(*) AS frequency
        FROM dataset_source
        WHERE NOT ({missing_condition_sql(definition.original_name)})
        GROUP BY 1
        ORDER BY frequency DESC, category ASC
    """
    counts_frame = connection.execute(counts_query).fetchdf()
    if counts_frame.empty:
        return []

    column_name = definition.normalized_name
    counts_frame = counts_frame.rename(columns={"category": column_name, "frequency": "count"})
    total_count = int(counts_frame["count"].sum())
    top_ten = counts_frame.head(10).copy()
    top_five = counts_frame.head(5).copy()
    other_count = max(total_count - int(top_five["count"].sum()), 0)
    if other_count:
        top_five = pd.concat(
            [
                top_five,
                pd.DataFrame([{column_name: "Other", "count": other_count}]),
            ],
            ignore_index=True,
        )

    bar_counts = top_ten.sort_values("count", ascending=True)
    bar_figure = px.bar(
        bar_counts,
        x="count",
        y=column_name,
        orientation="h",
        color="count",
        text="count",
        title=f"Top values in {column_name}",
        color_continuous_scale=["#DCCBFF", "#7C3AED"],
    )
    bar_figure.update_traces(textposition="outside", cliponaxis=False)
    style_figure(bar_figure)

    pie_figure = px.pie(
        top_five,
        names=column_name,
        values="count",
        title=f"Category share for {column_name}",
        hole=0.58,
        color_discrete_sequence=DASHBOARD_COLORS,
    )
    pie_figure.update_traces(
        textinfo="label+percent",
        pull=[0.06 if index == 0 else 0 for index in range(len(top_five))],
        marker={"line": {"color": "#fff8df", "width": 2}},
    )
    style_figure(pie_figure)

    return [
        build_chart_payload(
            chart_id=f"bar-{column_name}",
            chart_type="bar",
            title=f"Bar chart for {column_name}",
            description="Shows the most common categories using full-dataset counts.",
            figure=bar_figure,
        ),
        build_chart_payload(
            chart_id=f"pie-{column_name}",
            chart_type="pie",
            title=f"Pie chart for {column_name}",
            description="Shows the full-dataset category share for the top labels plus Other.",
            figure=pie_figure,
        ),
    ]


def build_exact_histogram_chart(
    connection: duckdb.DuckDBPyConnection,
    definition: ColumnDefinition,
    summary_lookup: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    summary = summary_lookup.get(definition.normalized_name, {})
    min_value = summary.get("min")
    max_value = summary.get("max")
    non_null_count = int(summary.get("non_null_count") or 0)
    if min_value is None or max_value is None or non_null_count <= 0:
        return None

    if float(min_value) == float(max_value):
        histogram_figure = go.Figure(
            data=[
                go.Bar(
                    x=[format_chart_tick(float(min_value))],
                    y=[non_null_count],
                    marker={"color": "#7C3AED", "line": {"color": "#F8E8AF", "width": 1}},
                )
            ]
        )
        histogram_figure.update_layout(
            title=f"Distribution of {definition.normalized_name}",
            xaxis_title=definition.normalized_name,
            yaxis_title="Count",
        )
        style_figure(histogram_figure)
        return build_chart_payload(
            chart_id=f"histogram-{definition.normalized_name}",
            chart_type="histogram",
            title=f"Histogram for {definition.normalized_name}",
            description="Displays the full-dataset distribution with exact bin counts.",
            figure=histogram_figure,
        )

    bin_count = min(30, max(10, min(non_null_count, 30)))
    bin_width = (float(max_value) - float(min_value)) / bin_count
    if bin_width <= 0:
        return None

    histogram_query = f"""
        WITH numeric AS (
            SELECT TRY_CAST({quote_identifier(definition.original_name)} AS DOUBLE) AS value
            FROM dataset_source
        ),
        filtered AS (
            SELECT value
            FROM numeric
            WHERE value IS NOT NULL
        ),
        bucketed AS (
            SELECT LEAST(
                {bin_count - 1},
                GREATEST(
                    0,
                    CAST(FLOOR((value - {to_sql_number(float(min_value))}) / {to_sql_number(bin_width)}) AS BIGINT)
                )
            ) AS bin_index
            FROM filtered
        )
        SELECT bin_index, COUNT(*) AS bin_count
        FROM bucketed
        GROUP BY 1
        ORDER BY 1
    """
    histogram_rows = connection.execute(histogram_query).fetchall()
    counts_by_bin = {int(row[0]): int(row[1]) for row in histogram_rows}
    labels: list[str] = []
    counts: list[int] = []
    for bin_index in range(bin_count):
        left_edge = float(min_value) + (bin_index * bin_width)
        right_edge = float(max_value) if bin_index == bin_count - 1 else left_edge + bin_width
        labels.append(f"{format_chart_tick(left_edge)} to {format_chart_tick(right_edge)}")
        counts.append(counts_by_bin.get(bin_index, 0))

    histogram_figure = go.Figure(
        data=[
            go.Bar(
                x=labels,
                y=counts,
                marker={"color": "#7C3AED", "line": {"color": "#F8E8AF", "width": 1}},
            )
        ]
    )
    histogram_figure.update_layout(
        title=f"Distribution of {definition.normalized_name}",
        xaxis_title=definition.normalized_name,
        yaxis_title="Count",
    )
    style_figure(histogram_figure)
    return build_chart_payload(
        chart_id=f"histogram-{definition.normalized_name}",
        chart_type="histogram",
        title=f"Histogram for {definition.normalized_name}",
        description="Displays the full-dataset distribution with exact bin counts.",
        figure=histogram_figure,
    )


def build_exact_box_chart(
    connection: duckdb.DuckDBPyConnection,
    definition: ColumnDefinition,
    summary_lookup: dict[str, dict[str, Any]],
    outlier_lookup: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    summary = summary_lookup.get(definition.normalized_name, {})
    q1_value = summary.get("q1")
    median_value = summary.get("median")
    q3_value = summary.get("q3")
    min_value = summary.get("min")
    max_value = summary.get("max")
    if None in {q1_value, median_value, q3_value, min_value, max_value}:
        return None

    lower_fence = float(min_value)
    upper_fence = float(max_value)
    outlier_summary = outlier_lookup.get(definition.normalized_name)
    if outlier_summary:
        whisker_query = f"""
            WITH numeric AS (
                SELECT TRY_CAST({quote_identifier(definition.original_name)} AS DOUBLE) AS value
                FROM dataset_source
            )
            SELECT
                MIN(value) FILTER (WHERE value >= {to_sql_number(float(outlier_summary["lower_bound"]))}) AS lower_fence,
                MAX(value) FILTER (WHERE value <= {to_sql_number(float(outlier_summary["upper_bound"]))}) AS upper_fence
            FROM numeric
            WHERE value IS NOT NULL
        """
        whisker_row = connection.execute(whisker_query).fetchone()
        if whisker_row:
            lower_fence = float(whisker_row[0]) if whisker_row[0] is not None else lower_fence
            upper_fence = float(whisker_row[1]) if whisker_row[1] is not None else upper_fence

    box_figure = go.Figure(
        data=[
            go.Box(
                name=definition.normalized_name,
                q1=[float(q1_value)],
                median=[float(median_value)],
                q3=[float(q3_value)],
                lowerfence=[lower_fence],
                upperfence=[upper_fence],
                boxpoints=False,
                fillcolor="rgba(20, 184, 166, 0.16)",
                marker={"color": "#7C3AED"},
                line={"color": "#14B8A6"},
            )
        ]
    )
    box_figure.update_layout(title=f"Outlier view for {definition.normalized_name}")
    style_figure(box_figure)
    return build_chart_payload(
        chart_id=f"box-{definition.normalized_name}",
        chart_type="box",
        title=f"Box plot for {definition.normalized_name}",
        description="Uses full-dataset quartiles and whiskers for an exact outlier summary.",
        figure=box_figure,
    )


def build_exact_density_chart(
    connection: duckdb.DuckDBPyConnection,
    x_definition: ColumnDefinition,
    y_definition: ColumnDefinition,
    summary_lookup: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    x_summary = summary_lookup.get(x_definition.normalized_name, {})
    y_summary = summary_lookup.get(y_definition.normalized_name, {})
    x_min = x_summary.get("min")
    x_max = x_summary.get("max")
    y_min = y_summary.get("min")
    y_max = y_summary.get("max")
    if None in {x_min, x_max, y_min, y_max}:
        return None

    x_min = float(x_min)
    x_max = float(x_max)
    y_min = float(y_min)
    y_max = float(y_max)
    if x_min == x_max or y_min == y_max:
        return None

    x_bin_width = (x_max - x_min) / MAX_EXACT_SCATTER_BINS
    y_bin_width = (y_max - y_min) / MAX_EXACT_SCATTER_BINS
    if x_bin_width <= 0 or y_bin_width <= 0:
        return None

    density_query = f"""
        WITH points AS (
            SELECT
                TRY_CAST({quote_identifier(x_definition.original_name)} AS DOUBLE) AS x_value,
                TRY_CAST({quote_identifier(y_definition.original_name)} AS DOUBLE) AS y_value
            FROM dataset_source
        ),
        filtered AS (
            SELECT x_value, y_value
            FROM points
            WHERE x_value IS NOT NULL AND y_value IS NOT NULL
        ),
        bucketed AS (
            SELECT
                LEAST(
                    {MAX_EXACT_SCATTER_BINS - 1},
                    GREATEST(
                        0,
                        CAST(FLOOR((x_value - {to_sql_number(x_min)}) / {to_sql_number(x_bin_width)}) AS BIGINT)
                    )
                ) AS x_bin,
                LEAST(
                    {MAX_EXACT_SCATTER_BINS - 1},
                    GREATEST(
                        0,
                        CAST(FLOOR((y_value - {to_sql_number(y_min)}) / {to_sql_number(y_bin_width)}) AS BIGINT)
                    )
                ) AS y_bin
            FROM filtered
        )
        SELECT x_bin, y_bin, COUNT(*) AS point_count
        FROM bucketed
        GROUP BY 1, 2
        ORDER BY 2, 1
    """
    density_rows = connection.execute(density_query).fetchall()
    if not density_rows:
        return None

    density_matrix = [
        [0 for _ in range(MAX_EXACT_SCATTER_BINS)]
        for _ in range(MAX_EXACT_SCATTER_BINS)
    ]
    for x_bin, y_bin, point_count in density_rows:
        density_matrix[int(y_bin)][int(x_bin)] = int(point_count)

    x_labels = [
        format_chart_tick(x_min + ((index + 0.5) * x_bin_width))
        for index in range(MAX_EXACT_SCATTER_BINS)
    ]
    y_labels = [
        format_chart_tick(y_min + ((index + 0.5) * y_bin_width))
        for index in range(MAX_EXACT_SCATTER_BINS)
    ]
    density_figure = go.Figure(
        data=[
            go.Heatmap(
                z=density_matrix,
                x=x_labels,
                y=y_labels,
                colorscale=HEATMAP_SCALE,
                hoverongaps=False,
            )
        ]
    )
    density_figure.update_layout(
        title=f"Density of {x_definition.normalized_name} vs {y_definition.normalized_name}",
        xaxis_title=x_definition.normalized_name,
        yaxis_title=y_definition.normalized_name,
    )
    style_figure(density_figure)
    return build_chart_payload(
        chart_id=f"density-{x_definition.normalized_name}-{y_definition.normalized_name}",
        chart_type="heatmap",
        title=f"Density map for {x_definition.normalized_name} and {y_definition.normalized_name}",
        description="Uses exact full-dataset density counts instead of a sampled scatter.",
        figure=density_figure,
    )


def build_exact_line_chart(
    connection: duckdb.DuckDBPyConnection,
    numeric_definition: ColumnDefinition,
    datetime_definition: ColumnDefinition | None,
) -> dict[str, Any] | None:
    numeric_identifier = quote_identifier(numeric_definition.original_name)
    if datetime_definition is not None:
        datetime_identifier = quote_identifier(datetime_definition.original_name)
        line_query = f"""
            WITH grouped AS (
                SELECT
                    TRY_CAST({datetime_identifier} AS TIMESTAMP) AS x_value,
                    AVG(TRY_CAST({numeric_identifier} AS DOUBLE)) AS y_value
                FROM dataset_source
                WHERE
                    TRY_CAST({datetime_identifier} AS TIMESTAMP) IS NOT NULL
                    AND TRY_CAST({numeric_identifier} AS DOUBLE) IS NOT NULL
                GROUP BY 1
            ),
            bucketed AS (
                SELECT
                    x_value,
                    y_value,
                    NTILE({MAX_EXACT_CHART_POINTS}) OVER (ORDER BY x_value) AS bucket
                FROM grouped
            )
            SELECT
                MIN(x_value) AS x_value,
                AVG(y_value) AS y_value
            FROM bucketed
            GROUP BY bucket
            ORDER BY x_value
        """
        line_frame = connection.execute(line_query).fetchdf()
        if line_frame.empty:
            return None

        line_figure = px.line(
            line_frame,
            x="x_value",
            y="y_value",
            markers=True,
            title=f"{numeric_definition.normalized_name} over {datetime_definition.normalized_name}",
            color_discrete_sequence=["#7C3AED"],
        )
    else:
        line_query = f"""
            WITH ordered AS (
                SELECT
                    ROW_NUMBER() OVER () AS record_index,
                    TRY_CAST({numeric_identifier} AS DOUBLE) AS y_value
                FROM dataset_source
            ),
            filtered AS (
                SELECT
                    record_index,
                    y_value,
                    NTILE({MAX_EXACT_CHART_POINTS}) OVER (ORDER BY record_index) AS bucket
                FROM ordered
                WHERE y_value IS NOT NULL
            )
            SELECT
                MIN(record_index) AS record_index,
                AVG(y_value) AS y_value
            FROM filtered
            GROUP BY bucket
            ORDER BY record_index
        """
        line_frame = connection.execute(line_query).fetchdf()
        if line_frame.empty:
            return None

        line_figure = px.line(
            line_frame,
            x="record_index",
            y="y_value",
            markers=True,
            title=f"{numeric_definition.normalized_name} across dataset order",
            color_discrete_sequence=["#7C3AED"],
        )

    line_figure.update_traces(
        line={"width": 3},
        marker={"size": 7, "color": "#F59E0B", "line": {"color": "#FFF5C3", "width": 1}},
        fill="tozeroy",
        fillcolor="rgba(124, 58, 237, 0.12)",
    )
    style_figure(line_figure)
    return build_chart_payload(
        chart_id=f"line-{numeric_definition.normalized_name}",
        chart_type="line",
        title=f"Line chart for {numeric_definition.normalized_name}",
        description="Uses exact full-dataset aggregation for the line trend.",
        figure=line_figure,
    )


def build_correlation_heatmap_chart(correlations: dict[str, Any]) -> dict[str, Any] | None:
    if not correlations.get("available"):
        return None

    matrix = correlations.get("matrix") or []
    columns = correlations.get("columns") or []
    if not matrix or not columns:
        return None

    heatmap_figure = go.Figure(
        data=[
            go.Heatmap(
                z=matrix,
                x=columns,
                y=columns,
                colorscale=HEATMAP_SCALE,
                zmin=-1,
                zmax=1,
                hoverongaps=False,
            )
        ]
    )
    heatmap_figure.update_layout(title="Correlation heatmap")
    style_figure(heatmap_figure)
    return build_chart_payload(
        chart_id="heatmap-correlations",
        chart_type="heatmap",
        title="Correlation heatmap",
        description="Highlights strong positive and negative relationships between numeric columns.",
        figure=heatmap_figure,
    )


def generate_exact_charts_for_cache_entry(cache_entry: AnalysisCacheEntry) -> list[dict[str, Any]]:
    payload = cache_entry.full_payload or cache_entry.preview_payload or {}
    summary_lookup = {
        item["column"]: item for item in payload.get("summary_statistics", [])
    }
    outlier_lookup = {
        item["column"]: item for item in payload.get("outliers", [])
    }

    with managed_dataset_connection(cache_entry) as connection:
        column_definitions = build_column_definitions_from_payload(payload, connection)
        numeric_definitions = [item for item in column_definitions if item.inferred_kind == "numeric"]
        datetime_definitions = [item for item in column_definitions if item.inferred_kind == "datetime"]
        categorical_definitions = [
            item for item in column_definitions if item.inferred_kind == "categorical"
        ]

        charts: list[dict[str, Any]] = []
        if categorical_definitions:
            charts.extend(build_exact_categorical_charts(connection, categorical_definitions[0]))

        if numeric_definitions:
            histogram_chart = build_exact_histogram_chart(connection, numeric_definitions[0], summary_lookup)
            if histogram_chart:
                charts.append(histogram_chart)

            box_chart = build_exact_box_chart(
                connection,
                numeric_definitions[0],
                summary_lookup,
                outlier_lookup,
            )
            if box_chart:
                charts.append(box_chart)

        if len(numeric_definitions) >= 2:
            density_chart = build_exact_density_chart(
                connection,
                numeric_definitions[0],
                numeric_definitions[1],
                summary_lookup,
            )
            if density_chart:
                charts.append(density_chart)

        if numeric_definitions:
            line_chart = build_exact_line_chart(
                connection,
                numeric_definitions[0],
                datetime_definitions[0] if datetime_definitions else None,
            )
            if line_chart:
                charts.append(line_chart)

    heatmap_chart = build_correlation_heatmap_chart(payload.get("correlations", {}))
    if heatmap_chart:
        charts.append(heatmap_chart)

    return sanitize_for_json(charts[: settings.max_chart_count])


def generate_charts_for_cache_entry(cache_entry: AnalysisCacheEntry) -> list[dict[str, Any]]:
    if cache_entry.processing_mode == "large":
        return generate_exact_charts_for_cache_entry(cache_entry)

    payload = cache_entry.full_payload or cache_entry.preview_payload or {}
    sample_frame = load_sample_for_cache_entry(cache_entry, limit=settings.chart_sample_rows)
    cleaned_frame, _ = clean_dataframe(sample_frame.copy())
    correlations = payload.get("correlations") or build_correlation_analysis(cleaned_frame)
    charts = generate_chart_specs(cleaned_frame, correlations)
    return sanitize_for_json(charts[: settings.max_chart_count])
def generate_transient_preview_charts(cache_entry: AnalysisCacheEntry) -> list[dict[str, Any]]:
    sample_frame = load_sample_for_cache_entry(cache_entry, limit=settings.chart_sample_rows)
    cleaned_frame, _ = clean_dataframe(sample_frame.copy())
    correlations = (cache_entry.preview_payload or {}).get("correlations") or build_correlation_analysis(cleaned_frame)
    return sanitize_for_json(generate_chart_specs(cleaned_frame, correlations)[: settings.max_chart_count])
def load_sample_for_cache_entry(cache_entry: AnalysisCacheEntry, *, limit: int) -> pd.DataFrame:
    parquet_path = Path(cache_entry.parquet_path) if cache_entry.parquet_path else None
    if parquet_path and parquet_path.exists():
        connection = duckdb.connect()
        try:
            return connection.execute(
                f"SELECT * FROM read_parquet({sql_literal(str(parquet_path))}) USING SAMPLE reservoir({limit} ROWS)"
            ).fetchdf()
        finally:
            connection.close()

    source_path = Path(cache_entry.storage_path)
    source_path = resolve_cache_entry_source_path(cache_entry)
    if cache_entry.file_type == "excel":
        return load_preview_sample(
            source_path=source_path,
            file_type=cache_entry.file_type,
            extension=source_path.suffix.lower(),
            limit=limit,
            sample=True,
        )
    connection = duckdb.connect()
    try:
        relation_sql = build_relation_sql_from_path(source_path, file_type=cache_entry.file_type)
        return connection.execute(f"SELECT * FROM {relation_sql} USING SAMPLE reservoir({limit} ROWS)").fetchdf()
    finally:
        connection.close()
def load_preview_sample(
    *,
    source_path: Path,
    file_type: str,
    extension: str,
    limit: int,
    sample: bool = False,
) -> pd.DataFrame:
    if file_type == "excel":
        if extension == ".xls":
            return load_legacy_excel_preview(source_path, limit, sample=sample)
        return load_modern_excel_preview(source_path, limit, sample=sample)

    connection = duckdb.connect()
    try:
        relation_sql = build_relation_sql_from_path(source_path, file_type=file_type)
        sample_clause = f" USING SAMPLE reservoir({limit} ROWS)" if sample else f" LIMIT {limit}"
        return connection.execute(f"SELECT * FROM {relation_sql}{sample_clause}").fetchdf()
    finally:
        connection.close()
def load_modern_excel_preview(source_path: Path, limit: int, *, sample: bool = False) -> pd.DataFrame:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return pd.DataFrame()
        columns = build_excel_columns(header_row)
        preview_rows = collect_excel_preview_rows(rows, columns, limit, sample=sample)
        return pd.DataFrame(preview_rows, columns=columns)
    finally:
        workbook.close()
def load_legacy_excel_preview(source_path: Path, limit: int, *, sample: bool = False) -> pd.DataFrame:
    rows_to_skip = 0
    sampled_rows: list[list[Any]] = []
    seen_rows = 0
    columns: list[str] | None = None
    rng = random.Random(SAMPLING_RANDOM_SEED)
    while True:
        chunk = pd.read_excel(
            source_path,
            nrows=settings.excel_chunk_rows if sample else limit,
            skiprows=(lambda row_index, offset=rows_to_skip: 0 < row_index <= offset) if rows_to_skip else None,
        )
        if chunk.empty:
            break
        if columns is None:
            columns = [str(column) for column in chunk.columns]
        chunk = chunk.reindex(columns=columns)
        row_values = [normalize_excel_row(row, len(columns)) for row in chunk.itertuples(index=False, name=None)]
        if sample:
            sampled_rows, seen_rows = apply_reservoir_sampling(sampled_rows, row_values, limit, seen_rows, rng)
        else:
            return pd.DataFrame(row_values[:limit], columns=columns)
        rows_to_skip += len(chunk)
        if len(chunk) < settings.excel_chunk_rows:
            break
    return pd.DataFrame(sampled_rows, columns=columns or [])
def collect_excel_preview_rows(
    rows: Any,
    columns: list[str],
    limit: int,
    *,
    sample: bool,
) -> list[list[Any]]:
    if not sample:
        preview_rows: list[list[Any]] = []
        for row in rows:
            preview_rows.append(normalize_excel_row(row, len(columns)))
            if len(preview_rows) >= limit:
                break
        return preview_rows
    sampled_rows: list[list[Any]] = []
    seen_rows = 0
    rng = random.Random(SAMPLING_RANDOM_SEED)
    for row in rows:
        sampled_rows, seen_rows = apply_reservoir_sampling(
            sampled_rows,
            [normalize_excel_row(row, len(columns))],
            limit,
            seen_rows,
            rng,
        )
    return sampled_rows
def normalize_excel_row(row: Any, width: int) -> list[Any]:
    padded_row = list(row[:width])
    if len(padded_row) < width:
        padded_row.extend([None] * (width - len(padded_row)))
    return padded_row
def apply_reservoir_sampling(
    sample_rows: list[list[Any]],
    incoming_rows: list[list[Any]],
    limit: int,
    seen_rows: int,
    rng: random.Random,
) -> tuple[list[list[Any]], int]:
    if limit <= 0:
        return sample_rows, seen_rows
    for row in incoming_rows:
        seen_rows += 1
        if len(sample_rows) < limit:
            sample_rows.append(row)
            continue
        replacement_index = rng.randint(1, seen_rows)
        if replacement_index <= limit:
            sample_rows[replacement_index - 1] = row
    return sample_rows, seen_rows
def load_sample_frame_from_connection(
    connection: duckdb.DuckDBPyConnection,
    limit: int,
) -> pd.DataFrame:
    return connection.execute(
        f"SELECT * FROM dataset_source USING SAMPLE reservoir({limit} ROWS)"
    ).fetchdf()
class managed_dataset_connection:
    def __init__(self, cache_entry: AnalysisCacheEntry):
        self.cache_entry = cache_entry
        self.connection: duckdb.DuckDBPyConnection | None = None
    def __enter__(self) -> duckdb.DuckDBPyConnection:
        self.connection = duckdb.connect()
        relation_sql = build_relation_sql_for_cache_entry(self.cache_entry)
        self.connection.execute(f"CREATE OR REPLACE TEMP VIEW dataset_source AS SELECT * FROM {relation_sql}")
        return self.connection
    def __exit__(self, exc_type, exc_value, traceback) -> None:
        if self.connection is not None:
            self.connection.close()
def build_relation_sql_for_cache_entry(cache_entry: AnalysisCacheEntry) -> str:
    if cache_entry.parquet_path:
        parquet_path = Path(cache_entry.parquet_path)
        if parquet_path.exists():
            return f"read_parquet({sql_literal(str(parquet_path))})"
    source_path = Path(cache_entry.storage_path)
    source_path = resolve_cache_entry_source_path(cache_entry)
    if cache_entry.file_type == "excel":
        raise ValueError("Excel datasets must be converted to Parquet before full analytics can run.")
    return build_relation_sql_from_path(source_path, file_type=cache_entry.file_type)
def build_relation_sql_from_path(source_path: Path, *, file_type: str) -> str:
    literal = sql_literal(str(source_path))
    if file_type == "csv":
        return f"read_csv_auto({literal}, header = true)"
    if file_type == "json":
        return f"read_json_auto({literal})"
    raise ValueError("Unsupported dataset source.")
def quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'
def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"
def missing_condition_sql(column_name: str) -> str:
    identifier = quote_identifier(column_name)
    return f"{identifier} IS NULL OR TRIM(CAST({identifier} AS VARCHAR)) = ''"
def resolve_cache_entry_source_path(cache_entry: AnalysisCacheEntry) -> Path:
    return ensure_local_storage_copy(
        storage_path=cache_entry.storage_path,
        storage_backend=cache_entry.storage_backend,
        storage_key=cache_entry.storage_key,
    )
def is_job_stale(cache_entry: AnalysisCacheEntry) -> bool:
    updated_at = cache_entry.updated_at
    if not updated_at:
        return True
    return updated_at < utcnow() - timedelta(seconds=settings.job_stale_after_seconds)
